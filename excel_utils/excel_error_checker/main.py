# TODO: 全体的に設計を見直す

from dataclasses import dataclass, field
from pathlib import Path
from typing import NamedTuple
from modules import logger
import pandas as pd

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

ERROR_CODES = ("#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A")
LOGGER = logger.create_logger()


class AddressIndex(NamedTuple):
    """0-indexのAddress情報を持つ"""

    row: int
    col: int


@dataclass
class Address:
    """A1, B2などの形式"""

    col: str
    row: str

    @property
    def to_str(self):
        return f"{self.col}{self.row}"


@dataclass
class Range:
    """A1:B2などの形式"""

    start: Address
    end: Address

    @property
    def to_str(self):
        return f"{self.start.to_str}:{self.end.to_str}"


@dataclass
class SheetChecker:
    df: pd.DataFrame
    is_checked_cells: list[list[bool]] = field(init=False)

    def __post_init__(self):
        mr, mc = self.df.shape
        self.is_checked_cells = [[False] * mc for _ in range(mr)]

    def search_error(self) -> tuple[Range, ...]:
        max_row, max_col = self.df.shape
        error_rngs: tuple[Range, ...] = ()
        for r in range(max_row):
            for c in range(max_col):
                if self.is_checked_cells[r][c]:
                    continue
                if self.df.iloc[r][c] in ERROR_CODES:
                    error_rngs += (self._get_error_range(r, c),)
                else:
                    self.is_checked_cells[r][c] = True
        return error_rngs

    def _get_error_range(self, start_row: int, start_col: int) -> Range:
        max_row, max_col = self.df.shape
        end_col = max_col - 1
        # TODO: 短縮化
        for now_row in range(start_row, max_row):
            if self.df.iloc[now_row][start_col] not in ERROR_CODES:
                break
            end_row = now_row
            for now_col in range(start_col, end_col + 1):
                if self.df.iloc[now_row][now_col] not in ERROR_CODES:
                    end_col = min(now_col - 1, end_col)
                    break
        start = AddressIndex(start_row, start_col)
        end = AddressIndex(end_row, end_col)
        self._update(start, end)
        return self._to_range(start, end)

    def _update(self, start: AddressIndex, end: AddressIndex) -> None:

        sr, sc = start
        er, ec = end
        for r in range(sr, er + 1):
            for c in range(sc, ec + 1):
                self.is_checked_cells[r][c] = True

    def _to_range(self, start: AddressIndex, end: AddressIndex) -> Range:

        return Range(self._to_address(start), self._to_address(end))

    def _to_address(self, address: AddressIndex) -> Address:
        r, c = address
        return Address(get_column_letter(c + 1), str(r + 1))


@dataclass
class FileChecker:
    file: Path
    wb: openpyxl.Workbook = field(init=False)

    def __post_init__(self):
        self.wb = load_workbook(self.file, read_only=True, data_only=True)

    def search_error(self) -> None:
        """対象ファイルにエラーが存在するか探す"""

        LOGGER.info(f"{self.file.name}のエラーチェックを行います")
        sht_to_error_rngs: dict[str, tuple[Range, ...]] = {}
        for sht in self.wb.sheetnames:
            df = pd.DataFrame(self.wb[sht].values)
            sht_to_error_rngs[sht] = SheetChecker(df).search_error()
        self._OutputLog(sht_to_error_rngs)

    def _OutputLog(self, sht_to_error_rngs: dict[str, tuple[Range, ...]]) -> None:
        """ログの出力を行う

        Args:
            sht_to_error_rngs (dict[str, tuple[Range, ...]]): シートと存在するエラー範囲の対応表
        """

        def formatted_rngs() -> list[str]:
            err = []
            for rng in error_rngs:
                err.append(rng.start.to_str if rng.start == rng.end else rng.to_str)
            return err

        for sht, error_rngs in sht_to_error_rngs.items():
            if not error_rngs:
                LOGGER.info(f"{sht}シートにエラーは存在しません")
                continue
            LOGGER.info(f"{sht}シートにエラーが存在します\n{', '.join(formatted_rngs())}")


def get_folder() -> list[Path]:
    """エラーチェックを行うファイルを取得する"""

    tgt_folder_path = Path(__file__).parent / "check_folder"
    return list(tgt_folder_path.glob("*.xlsx"))


def main():
    tgt_files = get_folder()
    for file in tgt_files:
        FileChecker(file).search_error()


if __name__ == "__main__":
    main()
